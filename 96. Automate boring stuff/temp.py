#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Dec 19 13:08:41 2020

@author: yukuo
"""

import os
import openpyxl

os.chdir('/Users/yukuo/Desktop/practice/96. Automate boring stuff/')

workbook = openpyxl.load_workbook('Book1.xlsx')

wb = openpyxl.Workbook()
print(wb.sheetnames)
wb['Sheet']['A1']=100
wb['Sheet']['A8']='你好'
wb.save('test.xlsx')

print(sh)